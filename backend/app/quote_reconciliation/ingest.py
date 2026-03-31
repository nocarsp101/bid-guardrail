# backend/app/quote_reconciliation/ingest.py
from __future__ import annotations

import csv
import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .aliases import ALIASES, CANONICAL_REQUIRED_FIELDS, CANONICAL_OPTIONAL_FIELDS


class IngestError(ValueError):
    def __init__(self, message: str, meta: Optional[Dict[str, Any]] = None):
        super().__init__(message)
        self.meta = meta or {}


NormalizedRow = Dict[str, Any]
Meta = Dict[str, Any]


def norm_header(h: Any) -> str:
    """
    Deterministic normalization for headers:
      - cast to str
      - strip
      - convert underscores to spaces (so unit_price -> UNIT PRICE)
      - collapse whitespace/newlines to single space
      - uppercase
    """
    if h is None:
        return ""
    s = str(h).replace("\n", " ").replace("\r", " ").replace("_", " ").strip()
    if not s:
        return ""
    s = " ".join(s.split())
    return s.upper()


def norm_item_code(v: Any) -> str:
    """
    Deterministic item normalization:
      - trim whitespace
      - strip leading zeros for purely numeric codes (e.g., 0010 -> 10)
      - keep non-numeric codes as-is (uppercased and space-collapsed)
    """
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    s = " ".join(s.split()).upper()

    # Strip leading zeros only if the whole token is digits
    if s.isdigit():
        s2 = s.lstrip("0")
        return s2 if s2 != "" else "0"
    return s


def to_num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def ingest_quote_lines(file_path: str) -> Tuple[List[NormalizedRow], Meta]:
    """
    Deterministic ingestion for QUOTE lines:
      - read CSV/XLSX (first sheet)
      - normalize headers (strip+uppercase+collapse whitespace/newlines)
      - map aliases EXACTLY
      - FAIL if missing required canonical field
      - FAIL if ambiguous (multiple cols map to same canonical field)
      - preserve item_raw exactly as read (for audit/proof)
      - normalize item code: trim + strip leading zeros
      - compute total deterministically if absent: qty * unit_price
    """
    ext = os.path.splitext(file_path.lower())[1]
    if ext == ".csv":
        raw_rows, headers = _read_csv(file_path)
    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raw_rows, headers = _read_xlsx_first_sheet(file_path)
    else:
        raise IngestError("Unsupported quote file type. Use CSV or XLSX.")

    mapping, mapping_meta = build_canonical_mapping(headers)

    meta: Meta = {
        "headers_detected": headers,
        "headers_normalized": [norm_header(h) for h in headers],
        "mapping_used": mapping,
        "mapping_alias_dict": ALIASES,
        "rows_raw_total": len(raw_rows),
        **mapping_meta,
        "normalization": {
            "headers": "strip + uppercase + collapse whitespace/newlines to single spaces",
            "item_raw": "original cell value as read (before normalization)",
            "item": "trim + strip leading zeros (numeric-only)",
        },
    }

    rows: List[NormalizedRow] = []
    for idx, r in enumerate(raw_rows):
        norm: NormalizedRow = {"_row_index": idx}

        for canonical, src_header in mapping.items():
            norm[canonical] = r.get(src_header)

        # preserve raw item exactly as read (for audit/proof)
        norm["item_raw"] = norm.get("item", "")

        # deterministic normalization
        norm["item"] = norm_item_code(norm.get("item"))

        qty = to_num(norm.get("qty"))
        up = to_num(norm.get("unit_price"))
        total = to_num(norm.get("total")) if "total" in norm else None

        if total is None and qty is not None and up is not None:
            norm["total"] = float(qty) * float(up)

        rows.append(norm)

    return rows, meta


def build_canonical_mapping(headers: List[str]) -> Tuple[Dict[str, str], Dict[str, Any]]:
    """
    Deterministic header alias mapping:
      - normalize headers
      - for each canonical field, find EXACT match from alias list
      - if required missing -> FAIL
      - if multiple headers map to same canonical -> FAIL
    """
    normalized_headers = [norm_header(h) for h in headers]

    norm_to_original: Dict[str, List[str]] = {}
    for orig, nh in zip(headers, normalized_headers):
        if nh == "":
            continue
        norm_to_original.setdefault(nh, []).append(orig)

    mapping: Dict[str, str] = {}
    ambiguous: Dict[str, List[str]] = {}
    missing: List[str] = []

    canonical_all = CANONICAL_REQUIRED_FIELDS + CANONICAL_OPTIONAL_FIELDS

    for canonical in canonical_all:
        alias_list = ALIASES.get(canonical, [])
        hits: List[str] = []
        for alias in alias_list:
            if alias in norm_to_original:
                hits.extend(norm_to_original[alias])

        if len(hits) > 1:
            ambiguous[canonical] = hits
            continue

        if len(hits) == 1:
            mapping[canonical] = hits[0]
        else:
            if canonical in CANONICAL_REQUIRED_FIELDS:
                missing.append(canonical)

    meta = {"mapping_missing": missing, "mapping_ambiguous": ambiguous}

    if ambiguous:
        raise IngestError("Ambiguous header mapping in QUOTE file (FAIL).", meta=meta)
    if missing:
        raise IngestError("Missing required canonical headers in QUOTE file (FAIL).", meta=meta)

    return mapping, meta


def _read_csv(path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = [row for row in reader]
    return rows, headers


def _read_xlsx_first_sheet(path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]

    rows: List[Dict[str, Any]] = []
    for r in rows_iter:
        row_dict: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row_dict[h] = r[i] if i < len(r) else None
        rows.append(row_dict)

    return rows, headers