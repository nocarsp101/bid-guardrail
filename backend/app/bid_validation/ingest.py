# backend/app/bid_validation/ingest.py
from __future__ import annotations

import csv
import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from .aliases import (
    ALIASES,
    CANONICAL_REQUIRED_FIELDS,
    CANONICAL_OPTIONAL_FIELDS,
)
from .normalize import norm_header, norm_item_code, to_bool, to_num


class IngestError(ValueError):
    def __init__(self, message: str, meta: Optional[Dict[str, Any]] = None):
        super().__init__(message)
        self.meta = meta or {}


NormalizedRow = Dict[str, Any]
Meta = Dict[str, Any]


def ingest_bid_items(file_path: str) -> Tuple[List[NormalizedRow], Meta]:
    """
    Deterministic ingestion:
      - read CSV/XLSX
      - deterministically find header row (XLSX)
      - normalize headers (strip + collapse whitespace + uppercase)
      - map aliases EXACTLY (no fuzzy / no inference)
      - FAIL if missing required canonical field
      - FAIL if ambiguous (multiple cols map to same canonical field)
      - normalize item codes: trim + strip leading zeros
      - if total missing/blank, compute total = qty * unit_price (deterministic)
      - skip summary/rollup rows deterministically
    """
    ext = os.path.splitext(file_path.lower())[1]
    if ext == ".csv":
        raw_rows, headers = _read_csv(file_path)
    elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raw_rows, headers, header_row_index = _read_xlsx_detect_header_row(file_path)
    else:
        raise IngestError("Unsupported bid items file type. Use CSV or XLSX.")

    mapping, mapping_meta = build_canonical_mapping(headers)

    meta: Meta = {
        "headers_detected": headers,
        "headers_normalized": [norm_header(h) for h in headers],
        "mapping_used": mapping,
        "mapping_alias_dict": ALIASES,
        "rows_raw_total": len(raw_rows),
        "rows_skipped_summary": 0,
        "normalization": {"headers": "strip + collapse whitespace + uppercase", "item": "trim + strip leading zeros"},
        **mapping_meta,
    }
    if ext != ".csv":
        meta["xlsx_header_row_index_0_based"] = header_row_index

    rows: List[NormalizedRow] = []
    for idx, r in enumerate(raw_rows):
        norm: NormalizedRow = {"_row_index": idx}

        for canonical, src_header in mapping.items():
            norm[canonical] = r.get(src_header)

        norm["item"] = norm_item_code(norm.get("item"))

        qty = to_num(norm.get("qty"))
        unit_price = to_num(norm.get("unit_price"))
        total = to_num(norm.get("total")) if "total" in norm else None

        if total is None and qty is not None and unit_price is not None:
            norm["total"] = float(qty) * float(unit_price)

        if _is_summary_row(norm):
            meta["rows_skipped_summary"] += 1
            continue
        
        # Skip deterministic placeholder rows (common in DOT exports)
        if _is_empty_placeholder_row(norm):
            meta["rows_skipped_summary"] += 1
            continue

        rows.append(norm)

    return rows, meta


def build_canonical_mapping(headers: List[str]) -> Tuple[Dict[str, str], Dict[str, Any]]:
    normalized_headers = [norm_header(h) for h in headers]

    norm_to_original: Dict[str, List[str]] = {}
    for orig, nh in zip(headers, normalized_headers):
        if nh == "":
            continue
        norm_to_original.setdefault(nh, []).append(orig)

    mapping: Dict[str, str] = {}
    ambiguous: Dict[str, List[str]] = {}
    missing: List[str] = []

    canonical_all = CANONICAL_REQUIRED_FIELDS + CANONICAL_OPTIONAL_FIELDS

    for canonical in canonical_all:
        alias_list = ALIASES.get(canonical, [])
        hits: List[str] = []
        for alias in alias_list:
            if alias in norm_to_original:
                hits.extend(norm_to_original[alias])

        if len(hits) > 1:
            ambiguous[canonical] = hits
            continue

        if len(hits) == 1:
            mapping[canonical] = hits[0]
        else:
            if canonical in CANONICAL_REQUIRED_FIELDS:
                missing.append(canonical)

    meta = {
        "mapping_missing": missing,
        "mapping_ambiguous": ambiguous,
        "mapping_used": mapping,
        "mapping_alias_dict": ALIASES,
        "normalization": {
            "headers": "strip + uppercase + collapse whitespace/newlines to single spaces",
            "item": "trim + strip leading zeros",
        },
        "headers_normalized": normalized_headers,
    }

    if ambiguous:
        raise IngestError(
            "Ambiguous header mapping: multiple columns map to same canonical field (FAIL).",
            meta=meta,
        )

    if missing:
        raise IngestError(
            "Missing required canonical headers (FAIL).",
            meta=meta,
        )

    return mapping, meta


def _read_csv(path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = [row for row in reader]
    return rows, headers


def _alias_universe() -> set[str]:
    """
    A set of all normalized alias strings (already uppercase) across all canonical fields.
    Used only to pick the header row deterministically for XLSX.
    """
    u: set[str] = set()
    for _, aliases in ALIASES.items():
        for a in aliases:
            u.add(a)
    return u


_ALIAS_UNIVERSE = _alias_universe()


def _score_header_row(cells: List[Any]) -> Tuple[int, int]:
    """
    Deterministic score:
      - hits = how many cells (normalized) are in alias universe
      - non_empty = how many non-empty cells
    """
    normed = [norm_header(c) for c in cells]
    non_empty = sum(1 for x in normed if x)
    hits = sum(1 for x in normed if x in _ALIAS_UNIVERSE)
    return hits, non_empty


def _read_xlsx_detect_header_row(path: str, scan_rows: int = 30) -> Tuple[List[Dict[str, Any]], List[str], int]:
    """
    Deterministically locate the header row in the first sheet:
      - scan first `scan_rows` rows
      - pick row with max (alias_hits, non_empty_cells)
      - require at least 2 alias hits to be considered a header row
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    # grab first N rows as lists
    raw_preview: List[List[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        raw_preview.append(list(row))
        if i + 1 >= scan_rows:
            break

    if not raw_preview:
        return [], [], 0

    best_idx = 0
    best_score = (-1, -1)
    for idx, row in enumerate(raw_preview):
        score = _score_header_row(row)
        if score > best_score:
            best_score = score
            best_idx = idx

    # require minimal confidence deterministically (no fuzzy):
    # if <2 alias hits, we treat first row as header (legacy behavior)
    alias_hits, _non_empty = best_score
    header_row_index = best_idx if alias_hits >= 2 else 0

    # Build headers from chosen header row
    header_row = raw_preview[header_row_index]
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    # Now read remaining rows AFTER header row
    rows: List[Dict[str, Any]] = []
    for r_idx, r in enumerate(ws.iter_rows(values_only=True)):
        if r_idx <= header_row_index:
            continue
        row_dict: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row_dict[h] = r[i] if i < len(r) else None
        rows.append(row_dict)

    return rows, headers, header_row_index


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _looks_like_item_number(s: str) -> bool:
    s = _s(s)
    if not s:
        return False
    return any(ch.isdigit() for ch in s)


def _is_summary_row(norm: Dict[str, Any]) -> bool:
    item = _s(norm.get("item"))
    desc = _s(norm.get("description"))
    notes = _s(norm.get("notes"))

    qty = to_num(norm.get("qty"))
    unit = _s(norm.get("unit"))
    unit_price = to_num(norm.get("unit_price"))
    total = to_num(norm.get("total"))
    excluded = to_bool(norm.get("excluded_flag"))

    blob = f"{item} {desc} {notes}".strip().lower()
    summary_tokens = ("grand total", "subtotal", "sub total", "bid total", "project total", "total")
    if any(tok in blob for tok in summary_tokens):
        if _looks_like_item_number(item):
            return False
        return True

    if (
        total is not None
        and not item
        and not desc
        and qty is None
        and not unit
        and unit_price is None
        and not notes
        and not excluded
    ):
        return True

    return False

def _is_empty_placeholder_row(norm: Dict[str, Any]) -> bool:
    """
    Deterministic: skip rows that look like placeholders/unpriced lines.
    Criteria:
      - item exists (often numbered)
      - qty missing/0
      - unit_price missing
      - total missing/0
      - description empty
    This prevents false FAILs on DOT exports that contain trailing empty numbered rows.
    """
    item = _s(norm.get("item"))
    desc = _s(norm.get("description"))

    qty = to_num(norm.get("qty"))
    unit_price = to_num(norm.get("unit_price"))
    total = to_num(norm.get("total"))

    if not item:
        return False

    qty_zeroish = (qty is None) or (abs(qty) < 1e-12)
    total_zeroish = (total is None) or (abs(total) < 1e-12)

    if qty_zeroish and unit_price is None and total_zeroish and desc == "":
        return True

    return False