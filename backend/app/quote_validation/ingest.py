# backend/app/quote_validation/ingest.py
from __future__ import annotations

import os
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from app.bid_validation.ingest import IngestError, build_canonical_mapping
from app.bid_validation.normalize import norm_item_code, to_num


QuoteRow = Dict[str, Any]
Meta = Dict[str, Any]


def ingest_quote_lines(file_path: str) -> Tuple[List[QuoteRow], Meta]:
    """
    Deterministic quote ingestion:
      - XLSX only (MVP)
      - canonical mapping via same alias dictionary
      - required: item, unit, qty, unit_price
      - normalize item (strip leading zeros)
      - compute extended line total deterministically: qty*unit_price
    """
    ext = os.path.splitext(file_path.lower())[1]
    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise IngestError("Unsupported quote file type. Use XLSX.")

    raw_rows, headers = _read_xlsx_first_sheet(file_path)
    mapping, mapping_meta = build_canonical_mapping(headers)

    meta: Meta = {
        "headers_detected": headers,
        "mapping_used": mapping,
        **mapping_meta,
        "rows_raw_total": len(raw_rows),
    }

    rows: List[QuoteRow] = []
    for idx, r in enumerate(raw_rows):
        norm: QuoteRow = {"_row_index": idx}

        for canonical, src_header in mapping.items():
            norm[canonical] = r.get(src_header)

        norm["item"] = norm_item_code(norm.get("item"))
        qty = to_num(norm.get("qty"))
        unit_price = to_num(norm.get("unit_price"))
        norm["computed_total"] = (float(qty) * float(unit_price)) if (qty is not None and unit_price is not None) else None

        rows.append(norm)

    return rows, meta


def _read_xlsx_first_sheet(path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: List[Dict[str, Any]] = []

    for r in rows_iter:
        row_dict: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            row_dict[h] = r[i] if i < len(r) else None
        rows.append(row_dict)

    return rows, headers